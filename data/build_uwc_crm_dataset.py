"""
Build UWC CRM sample dataset — Excel workbook with 18 module sheets,
10+ records each, FK-linked, ready for Zoho CRM import.

Output: UWC_CRM_Sample_Dataset.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from datetime import date

# ---------- THEME ----------
UWC_NAVY = "003087"
UWC_RED = "C8102E"
ZOHO_BLUE = "3B6EF5"
LIGHT_GREY = "F1F5F9"
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", start_color=UWC_NAVY)
BODY_FONT = Font(name="Arial", size=10)
BODY_ALT_FILL = PatternFill("solid", start_color="F8FAFC")
ID_FONT = Font(name="Consolas", size=10, color="2563EB")
THIN = Side(style="thin", color="E2E8F0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP = Alignment(wrap_text=True, vertical="top", horizontal="left")


def style_sheet(ws, n_rows, n_cols, col_widths=None, freeze="A2"):
    """Apply header style + alternating row fill + borders + widths."""
    # Header row
    for c in range(1, n_cols + 1):
        cell = ws.cell(1, c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[1].height = 32
    # Body rows
    for r in range(2, n_rows + 2):
        for c in range(1, n_cols + 1):
            cell = ws.cell(r, c)
            cell.font = BODY_FONT
            cell.alignment = WRAP_TOP
            cell.border = BORDER
            if r % 2 == 0:
                cell.fill = BODY_ALT_FILL
            # ID column styling
            if c == 1:
                cell.font = ID_FONT
    # Widths
    if col_widths:
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = freeze


def write_sheet(wb, sheet_name, headers, rows, col_widths=None):
    ws = wb.create_sheet(sheet_name)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    style_sheet(ws, len(rows), len(headers), col_widths)
    return ws


# ---------- BUILD ----------
wb = Workbook()
default = wb.active
wb.remove(default)


# ============================================================
# SHEET 0: README — Schema overview + FK map
# ============================================================
readme = wb.create_sheet("README")
readme["A1"] = "UWC International — CRM Sample Dataset"
readme["A1"].font = Font(name="Arial", size=22, bold=True, color=UWC_NAVY)
readme.merge_cells("A1:F1")

readme["A2"] = "Demo data for the Zoho CRM and Applications Platform tender (ref 5128). 18 modules, 200+ records, fully FK-linked."
readme["A2"].font = Font(name="Arial", size=11, italic=True, color="475569")
readme.merge_cells("A2:F2")

readme["A4"] = "Generated"
readme["B4"] = str(date.today())
readme["A5"] = "Prepared by"
readme["B5"] = "A2Z Cloud (Pranesh Subramani)"
readme["A6"] = "Target"
readme["B6"] = "Zoho CRM EU DC (crm.zoho.eu, org 20114943111)"
readme["A7"] = "Import order"
readme["B7"] = "Top-down by sheet order — parents first (National_Committees → Programmes → Accounts → Contacts → Leads → Applications → child records)"
for r in range(4, 8):
    readme.cell(r, 1).font = Font(name="Arial", size=10, bold=True, color="475569")

readme["A9"] = "Module overview"
readme["A9"].font = Font(name="Arial", size=14, bold=True, color=UWC_NAVY)

modules_overview = [
    ("Sheet", "Module", "Records", "Domain", "Key FK references", "Notes"),
    ("01", "National_Committees", 12, "Governance", "—", "Top-of-hierarchy — every other record traces back to an NC"),
    ("02", "Programmes", 18, "Selection", "—", "18 UWC colleges (the Two-Year IB Diploma destinations)"),
    ("03", "Accounts", 12, "Comms / Volunteer", "—", "Schools + Donor Organisations (Zoho Accounts module)"),
    ("04", "Contacts", 30, "People", "NC, Account", "Students, Parents, Alumni, Selectors, Donors (5 sub-types)"),
    ("05", "Leads", 12, "App Core", "NC", "Pre-application enquiries (convert to Contact + Application)"),
    ("06", "Applications", 13, "App Core", "Contact, NC, Programme", "The pipeline record (Zoho Deals analogue) — Blueprint-managed"),
    ("07", "Interviews", 12, "Selection", "Application, Contact, NC", "1-to-many on Application — interview slots + outcomes"),
    ("08", "Review_Scores", 18, "Selection", "Application, Contact (reviewer)", "Subform on Application — 2 reviewers per app, blind scoring"),
    ("09", "Documents", 18, "App Core", "Application, Contact", "Attachment metadata (transcript, passport, recommendation letters)"),
    ("10", "Consent_Records", 12, "Audit", "Contact, Application", "GDPR + safeguarding consent audit trail"),
    ("11", "Financial_Aid", 12, "Selection", "Application, Contact", "Scholarship review records — finance team workflow"),
    ("12", "Tasks", 15, "Comms", "Application, Contact (owner)", "To-dos across the team — due dates + priority"),
    ("13", "Meetings", 12, "Comms", "Application, Contact, NC", "Selection committee + interview + governance meetings"),
    ("14", "Calls", 12, "Comms", "Application, Contact", "Phone follow-ups (parent calls, interview confirmations)"),
    ("15", "Campaigns", 10, "Selection", "NC, Programme", "Selection campaigns + outreach (paid/organic + events)"),
    ("16", "Stage_History", 18, "Audit", "Application", "Audit trail of every stage transition on every Application"),
    ("17", "Training_Records", 12, "Volunteer", "Contact (Selector)", "Selector / reviewer training completion records"),
    ("18", "Audit_Log", 15, "Audit", "any record", "System-wide audit — unlock events, GDPR access, role changes"),
]

start_row = 10
for r_idx, row in enumerate(modules_overview, start=start_row):
    for c_idx, val in enumerate(row, start=1):
        c = readme.cell(r_idx, c_idx, val)
        if r_idx == start_row:
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        else:
            c.font = BODY_FONT
            c.alignment = WRAP_TOP
            if r_idx % 2 == 0:
                c.fill = BODY_ALT_FILL
        c.border = BORDER

for col, w in zip("ABCDEF", [8, 22, 11, 18, 38, 56]):
    readme.column_dimensions[col].width = w
readme.row_dimensions[start_row].height = 28

# Conventions section
conv_row = start_row + len(modules_overview) + 2
readme.cell(conv_row, 1, "ID conventions").font = Font(name="Arial", size=14, bold=True, color=UWC_NAVY)
conventions = [
    ("Prefix", "Example", "Meaning"),
    ("NC-XX", "NC-DE", "National Committee, 2-letter ISO country code"),
    ("PRG-NNN", "PRG-001", "Programme record (UWC college)"),
    ("PRG-CMP-XXNN", "PRG-CMP-DE26", "NC selection campaign — NC code + 2-digit year"),
    ("ACC-TYPE-NNN", "ACC-SCH-001", "Account: SCH=school, DON=donor organisation, VND=vendor"),
    ("CON-TYPE-NNN", "CON-STU-001", "Contact: STU=student, PAR=parent, ALM=alumnus, SEL=selector, DON=donor"),
    ("LEAD-NNN", "LEAD-001", "Lead (pre-application enquiry)"),
    ("APP-YYYY-NNNNN", "APP-2026-00073", "Application — year + 5-digit sequence"),
    ("INT-NNN", "INT-001", "Interview slot"),
    ("REV-NNN", "REV-001", "Review score record"),
    ("DOC-NNN", "DOC-001", "Document attachment"),
    ("CR-NNN", "CR-001", "Consent record"),
    ("FA-NNN", "FA-001", "Financial Aid review"),
    ("TSK-NNN", "TSK-001", "Task"),
    ("MTG-NNN", "MTG-001", "Meeting"),
    ("CALL-NNN", "CALL-001", "Call"),
    ("CMP-NNN", "CMP-001", "Marketing campaign"),
    ("SH-NNN", "SH-001", "Stage History audit entry"),
    ("TRN-NNN", "TRN-001", "Training record"),
    ("AUD-NNN", "AUD-001", "Audit log entry"),
]
for ridx, row in enumerate(conventions, start=conv_row + 1):
    for cidx, val in enumerate(row, start=1):
        c = readme.cell(ridx, cidx, val)
        if ridx == conv_row + 1:
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
        else:
            c.font = BODY_FONT
            if ridx % 2 == 0:
                c.fill = BODY_ALT_FILL
        c.border = BORDER
        c.alignment = WRAP_TOP


# ============================================================
# SHEET 1: National_Committees (12)
# ============================================================
nc_headers = ["NC ID", "NC Name", "Country", "ISO Code", "Region", "Director Contact", "Email",
              "Phone", "Office Address", "Established", "Active Programmes", "Status"]
nc_rows = [
    ["NC-DE", "UWC Deutschland", "Germany", "DE", "Europe", "Dr. Marcus Weber", "marcus.weber@de.uwc.org",
     "+49 228 9657 100", "Bornheimer Str. 80, 53111 Bonn, Germany", 1962, 3, "Active"],
    ["NC-BR", "UWC Brazil", "Brazil", "BR", "Americas", "Ana Carvalho", "ana.carvalho@br.uwc.org",
     "+55 11 3151 4000", "Rua Augusta 1240, São Paulo SP 01304-001", 1971, 2, "Active"],
    ["NC-IN", "UWC India", "India", "IN", "Asia", "Vikram Mehta", "vikram.mehta@in.uwc.org",
     "+91 22 6770 9000", "Worli Sea Face, Mumbai 400018", 1980, 3, "Active"],
    ["NC-KE", "UWC Kenya", "Kenya", "KE", "Africa", "Joyce Wambui", "joyce.wambui@ke.uwc.org",
     "+254 20 4444 200", "Lavington, Nairobi", 1976, 2, "Active"],
    ["NC-UK", "UWC United Kingdom", "United Kingdom", "GB", "Europe", "James Whitfield OBE", "james.whitfield@uk.uwc.org",
     "+44 20 7269 7800", "55 New Oxford Street, London WC1A 1BS", 1962, 4, "Active"],
    ["NC-US", "UWC USA", "United States", "US", "Americas", "Sarah O'Brien", "sarah.obrien@usa.uwc.org",
     "+1 575 454 4200", "PO Box 248, Montezuma NM 87731", 1982, 4, "Active"],
    ["NC-JP", "UWC Japan", "Japan", "JP", "Asia", "Hiroshi Tanaka", "hiroshi.tanaka@jp.uwc.org",
     "+81 3 5544 5588", "Akasaka 6-13-7, Minato-ku, Tokyo 107-0052", 1974, 3, "Active"],
    ["NC-NO", "UWC Norge", "Norway", "NO", "Europe", "Ingrid Olafsson", "ingrid.olafsson@no.uwc.org",
     "+47 22 56 6700", "Pilestredet 32, 0166 Oslo", 1962, 2, "Active"],
    ["NC-IT", "UWC Italia", "Italy", "IT", "Europe", "Giulia Romano", "giulia.romano@it.uwc.org",
     "+39 06 808 4500", "Via del Boschetto 14, 00184 Roma", 1972, 2, "Active"],
    ["NC-FR", "UWC France", "France", "FR", "Europe", "Pierre Lefebvre", "pierre.lefebvre@fr.uwc.org",
     "+33 1 4544 3300", "12 rue de la Sorbonne, 75005 Paris", 1971, 2, "Active"],
    ["NC-AR", "UWC Argentina", "Argentina", "AR", "Americas", "Mariana Acosta", "mariana.acosta@ar.uwc.org",
     "+54 11 4811 2200", "Av. del Libertador 1235, Buenos Aires", 1989, 1, "Active"],
    ["NC-CA", "UWC Canada", "Canada", "CA", "Americas", "Margaret Robinson", "margaret.robinson@ca.uwc.org",
     "+1 250 391 2411", "650 Pearson College Dr, Victoria BC V9C 4H7", 1974, 2, "Active"],
]
write_sheet(wb, "01_National_Committees", nc_headers, nc_rows,
            [9, 24, 16, 9, 11, 22, 30, 18, 36, 12, 17, 9])


# ============================================================
# SHEET 2: Programmes (18 UWC colleges + 8 selection campaigns = 26)
# ============================================================
prg_headers = ["Programme ID", "Programme Name", "Programme Type", "Host Country", "Location",
               "Start Date", "End Date", "Capacity", "Tuition (GBP, 2yr)", "Scholarship %",
               "Owning NC", "Status"]
prg_rows = [
    ["PRG-001", "UWC Atlantic", "Two-Year IB Diploma", "United Kingdom", "St. Donat's Castle, Llantwit Major, Wales",
     "2026-08-25", "2028-05-30", 350, 90000, 75, "NC-UK", "Active"],
    ["PRG-002", "UWC Adriatic", "Two-Year IB Diploma", "Italy", "Duino, Trieste",
     "2026-09-01", "2028-06-15", 200, 72000, 70, "NC-IT", "Active"],
    ["PRG-003", "UWC Costa Rica", "Two-Year IB Diploma", "Costa Rica", "Santa Ana, San José",
     "2026-08-15", "2028-05-30", 196, 60000, 80, "NC-US", "Active"],
    ["PRG-004", "UWC Dilijan", "Two-Year IB Diploma", "Armenia", "Dilijan",
     "2026-08-20", "2028-06-10", 230, 84000, 65, "NC-IT", "Active"],
    ["PRG-005", "UWC East Africa", "Two-Year IB Diploma", "Tanzania", "Moshi",
     "2026-08-10", "2028-05-25", 180, 56000, 85, "NC-KE", "Active"],
    ["PRG-006", "UWC ISAK Japan", "Two-Year IB Diploma", "Japan", "Karuizawa, Nagano",
     "2026-09-01", "2028-06-15", 200, 96000, 70, "NC-JP", "Active"],
    ["PRG-007", "UWC Maastricht", "Two-Year IB Diploma", "Netherlands", "Maastricht",
     "2026-08-20", "2028-06-30", 920, 48000, 50, "NC-UK", "Active"],
    ["PRG-008", "UWC Mahindra College", "Two-Year IB Diploma", "India", "Paud, Pune Maharashtra",
     "2026-07-25", "2028-05-20", 252, 50000, 75, "NC-IN", "Active"],
    ["PRG-009", "UWC Mostar", "Two-Year IB Diploma", "Bosnia and Herzegovina", "Mostar",
     "2026-09-05", "2028-06-20", 220, 52000, 80, "NC-IT", "Active"],
    ["PRG-010", "Pearson College UWC", "Two-Year IB Diploma", "Canada", "Victoria, British Columbia",
     "2026-08-22", "2028-05-31", 200, 88000, 70, "NC-CA", "Active"],
    ["PRG-011", "UWC Red Cross Nordic", "Two-Year IB Diploma", "Norway", "Flekke, Fjaler",
     "2026-08-25", "2028-06-10", 200, 78000, 75, "NC-NO", "Active"],
    ["PRG-012", "UWC Robert Bosch College", "Two-Year IB Diploma", "Germany", "Freiburg im Breisgau",
     "2026-08-20", "2028-06-30", 200, 78000, 70, "NC-DE", "Active"],
    ["PRG-013", "UWC Changshu China", "Two-Year IB Diploma", "China", "Changshu, Jiangsu",
     "2026-08-25", "2028-06-15", 552, 64000, 60, "NC-IN", "Active"],
    ["PRG-014", "UWC Thailand", "Two-Year IB Diploma", "Thailand", "Phuket",
     "2026-08-10", "2028-06-15", 700, 60000, 55, "NC-JP", "Active"],
    ["PRG-015", "UWC South East Asia (Singapore)", "Two-Year IB Diploma", "Singapore", "Dover + East campuses",
     "2026-08-12", "2028-06-15", 5000, 80000, 30, "NC-IN", "Active"],
    ["PRG-016", "UWC-USA", "Two-Year IB Diploma", "United States", "Montezuma, New Mexico",
     "2026-08-15", "2028-06-01", 235, 92000, 85, "NC-US", "Active"],
    ["PRG-017", "Waterford Kamhlaba UWC", "Two-Year IB Diploma", "Eswatini", "Mbabane",
     "2026-08-10", "2028-06-10", 600, 42000, 80, "NC-KE", "Active"],
    ["PRG-018", "Li Po Chun UWC", "Two-Year IB Diploma", "Hong Kong", "Wu Kai Sha, Ma On Shan",
     "2026-08-22", "2028-06-15", 268, 82000, 65, "NC-IN", "Active"],
    # NC selection campaigns
    ["PRG-CMP-DE26", "UWC Deutschland 2026 Auswahl", "NC Selection Campaign", "Germany", "Bonn (online + in-person)",
     "2025-10-01", "2026-06-30", 25, 0, 100, "NC-DE", "Active"],
    ["PRG-CMP-BR26", "UWC Brazil 2026 Selection", "NC Selection Campaign", "Brazil", "São Paulo, Rio de Janeiro",
     "2025-10-01", "2026-06-30", 20, 0, 100, "NC-BR", "Active"],
    ["PRG-CMP-IN26", "UWC India 2026 Selection", "NC Selection Campaign", "India", "Mumbai + Delhi + Bengaluru",
     "2025-10-01", "2026-06-30", 35, 0, 100, "NC-IN", "Active"],
    ["PRG-CMP-KE26", "UWC Kenya 2026 Selection", "NC Selection Campaign", "Kenya", "Nairobi",
     "2025-10-01", "2026-06-30", 15, 0, 100, "NC-KE", "Active"],
    ["PRG-CMP-UK26", "UWC United Kingdom 2026 Selection", "NC Selection Campaign", "United Kingdom", "London + regional centres",
     "2025-10-01", "2026-06-30", 28, 0, 100, "NC-UK", "Active"],
    ["PRG-CMP-US26", "UWC USA 2026 Selection (Davis Programme)", "NC Selection Campaign", "United States", "Montezuma + virtual",
     "2025-10-01", "2026-06-30", 40, 0, 100, "NC-US", "Active"],
    ["PRG-CMP-JP26", "UWC Japan 2026 Senbatsu", "NC Selection Campaign", "Japan", "Tokyo",
     "2025-10-01", "2026-06-30", 18, 0, 100, "NC-JP", "Active"],
    ["PRG-CMP-NO26", "UWC Norge 2026 Utvalg", "NC Selection Campaign", "Norway", "Oslo + Trondheim",
     "2025-10-01", "2026-06-30", 14, 0, 100, "NC-NO", "Active"],
]
write_sheet(wb, "02_Programmes", prg_headers, prg_rows,
            [16, 36, 22, 16, 36, 12, 12, 9, 14, 10, 10, 9])


# ============================================================
# SHEET 3: Accounts (12 — schools + donor orgs)
# ============================================================
acc_headers = ["Account ID", "Account Name", "Account Type", "Country", "Primary Contact",
               "Email", "Phone", "Website", "Annual Revenue (GBP)", "Owning NC", "Tags", "Status"]
acc_rows = [
    ["ACC-SCH-001", "Gymnasium am Wirteltor", "Feeder School", "Germany", "Dr. Heinrich Klein",
     "h.klein@gym-wirteltor.de", "+49 241 4775 0", "https://gym-wirteltor.de", 0, "NC-DE", "feeder; abitur", "Active"],
    ["ACC-SCH-002", "Colégio São Luís", "Feeder School", "Brazil", "Prof. Carlos Almeida",
     "c.almeida@saoluis.com.br", "+55 11 2371 1200", "https://saoluis.com.br", 0, "NC-BR", "feeder; jesuit", "Active"],
    ["ACC-SCH-003", "Cathedral & John Connon School", "Feeder School", "India", "Meera Isaac",
     "head@cathedral.org.in", "+91 22 2204 1670", "https://cathedral.org.in", 0, "NC-IN", "feeder; mumbai", "Active"],
    ["ACC-SCH-004", "Brookhouse School Nairobi", "Feeder School", "Kenya", "Patrick Mwangi",
     "p.mwangi@brookhouse.ac.ke", "+254 20 6605 100", "https://brookhouse.ac.ke", 0, "NC-KE", "feeder; ib", "Active"],
    ["ACC-SCH-005", "Westminster School", "Feeder School", "United Kingdom", "Dr. Gary Savage",
     "headmaster@westminster.org.uk", "+44 20 7963 1000", "https://westminster.org.uk", 0, "NC-UK", "feeder; oxbridge", "Active"],
    ["ACC-SCH-006", "Phillips Academy Andover", "Feeder School", "United States", "Dr. Raynard Kington",
     "headofschool@andover.edu", "+1 978 749 4000", "https://andover.edu", 0, "NC-US", "feeder; davis programme", "Active"],
    ["ACC-DON-001", "Davis United World College Scholars Programme", "Donor Organisation", "United States", "Philip Geier Jr.",
     "info@davisuwcscholars.org", "+1 802 333 2030", "https://davisuwcscholars.org", 45000000, "NC-US", "donor; tier-1", "Active"],
    ["ACC-DON-002", "Shelby Davis Foundation", "Donor Organisation", "United States", "Shelby M.C. Davis",
     "office@davisfoundation.org", "+1 802 658 8200", "https://davisfoundation.org", 28000000, "NC-US", "donor; tier-1; founder", "Active"],
    ["ACC-DON-003", "Robert Bosch Stiftung", "Donor Organisation", "Germany", "Bernhard Straub",
     "info@bosch-stiftung.de", "+49 711 46084 0", "https://bosch-stiftung.de", 18000000, "NC-DE", "donor; tier-1; corporate", "Active"],
    ["ACC-DON-004", "Mahindra Group Foundation", "Donor Organisation", "India", "Anand Mahindra",
     "csr@mahindra.com", "+91 22 2490 1441", "https://mahindra.com/csr", 12000000, "NC-IN", "donor; tier-1; corporate", "Active"],
    ["ACC-DON-005", "The Sasakawa Peace Foundation", "Donor Organisation", "Japan", "Nobuko Hori",
     "info@spf.org", "+81 3 5157 5430", "https://spf.org", 8500000, "NC-JP", "donor; tier-2; peace", "Active"],
    ["ACC-VND-001", "Zoho Corporation", "Vendor", "India", "Sridhar Vembu",
     "uwc-enterprise@zohocorp.com", "+91 44 7155 4444", "https://zoho.com", 0, "NC-IN", "vendor; crm", "Active"],
]
write_sheet(wb, "03_Accounts", acc_headers, acc_rows,
            [13, 36, 18, 16, 22, 32, 18, 32, 14, 10, 24, 9])


# ============================================================
# SHEET 4: Contacts (30 — Students, Parents, Alumni, Selectors, Donors)
# ============================================================
con_headers = ["Contact ID", "Full Name", "Type", "Date of Birth", "Gender", "Nationality",
               "Email", "Mobile", "City", "Country", "Owning NC", "Related Account",
               "Notes", "Status"]
con_rows = [
    # Students
    ["CON-STU-001", "Lena Fischer", "Student", "2009-04-12", "Female", "German",
     "lena.fischer@noemail.invalid", "+49 221 5555 1234", "Cologne", "Germany", "NC-DE", "ACC-SCH-001",
     "Abitur projected 1.2 · school council president", "Active"],
    ["CON-STU-002", "Klaus Berger", "Student", "2009-08-22", "Male", "German",
     "klaus.berger@noemail.invalid", "+49 30 5555 4321", "Berlin", "Germany", "NC-DE", "ACC-SCH-001",
     "Maths Olympiad national finalist", "Active"],
    ["CON-STU-003", "Anneke Müller", "Student", "2010-01-15", "Female", "German",
     "anneke.muller@noemail.invalid", "+49 89 5555 8800", "Munich", "Germany", "NC-DE", "ACC-SCH-001",
     "Debate champion · climate activist", "Active"],
    ["CON-STU-004", "Amara Diallo", "Student", "2009-11-03", "Female", "German (Senegalese heritage)",
     "amara.diallo@noemail.invalid", "+49 69 5555 6677", "Frankfurt", "Germany", "NC-DE", "ACC-SCH-001",
     "MUN ambassador · Wolof + French native", "Active"],
    ["CON-STU-005", "Sofia Almeida", "Student", "2009-08-15", "Female", "Brazilian",
     "sofia.almeida@noemail.invalid", "+55 11 9 4567 1234", "São Paulo", "Brazil", "NC-BR", "ACC-SCH-002",
     "Public-school scholarship · favela community leader", "Active"],
    ["CON-STU-006", "Rafael Mendes", "Student", "2010-02-28", "Male", "Brazilian",
     "rafael.mendes@noemail.invalid", "+55 21 9 8765 4321", "Rio de Janeiro", "Brazil", "NC-BR", "ACC-SCH-002",
     "Mathematics Olympiad medallist", "Active"],
    ["CON-STU-007", "Aarav Sharma", "Student", "2009-06-10", "Male", "Indian",
     "aarav.sharma@noemail.invalid", "+91 98765 43210", "Mumbai", "India", "NC-IN", "ACC-SCH-003",
     "Robotics team captain", "Active"],
    ["CON-STU-008", "Priya Iyer", "Student", "2010-04-25", "Female", "Indian",
     "priya.iyer@noemail.invalid", "+91 98765 11122", "Bengaluru", "India", "NC-IN", "ACC-SCH-003",
     "Bharatanatyam dancer · Bharat scholar", "Active"],
    ["CON-STU-009", "Kamau Njoroge", "Student", "2009-09-19", "Male", "Kenyan",
     "kamau.njoroge@noemail.invalid", "+254 722 444 555", "Nairobi", "Kenya", "NC-KE", "ACC-SCH-004",
     "Wildlife conservation volunteer · Maasai Mara project", "Active"],
    ["CON-STU-010", "Yuki Tanaka", "Student", "2010-03-08", "Female", "Japanese",
     "yuki.tanaka@noemail.invalid", "+81 90 1234 5678", "Tokyo", "Japan", "NC-JP", "",
     "Hosei girls' school · global issues club president", "Active"],
    ["CON-STU-011", "Oliver Whitfield", "Student", "2009-12-01", "Male", "British",
     "o.whitfield@noemail.invalid", "+44 7700 900 123", "London", "United Kingdom", "NC-UK", "ACC-SCH-005",
     "Cambridge offer-holder applying via UWC route", "Active"],
    ["CON-STU-012", "Ayşe Demir", "Student", "2009-10-17", "Female", "Turkish",
     "ayse.demir@noemail.invalid", "+90 532 555 1234", "Istanbul", "Turkey", "NC-IT", "",
     "Robotics, refugee education volunteer (Syrian community)", "Active"],
    # Parents / Guardians
    ["CON-PAR-001", "Petra Fischer", "Parent", "1973-05-22", "Female", "German",
     "petra.fischer@example.de", "+49 221 5555 1234", "Cologne", "Germany", "NC-DE", "",
     "Mother of Lena Fischer · single parent · social worker", "Active"],
    ["CON-PAR-002", "Maria Almeida", "Parent", "1971-11-04", "Female", "Brazilian",
     "maria.almeida@example.com.br", "+55 11 9 8888 2222", "São Paulo", "Brazil", "NC-BR", "",
     "Mother of Sofia Almeida · public health nurse", "Active"],
    ["CON-PAR-003", "Anand Sharma", "Parent", "1972-07-15", "Male", "Indian",
     "anand.sharma@example.in", "+91 98765 99888", "Mumbai", "India", "NC-IN", "",
     "Father of Aarav · chartered accountant", "Active"],
    ["CON-PAR-004", "Grace Njoroge", "Parent", "1975-03-30", "Female", "Kenyan",
     "grace.njoroge@example.co.ke", "+254 722 777 888", "Nairobi", "Kenya", "NC-KE", "",
     "Mother of Kamau · primary school head teacher", "Active"],
    ["CON-PAR-005", "Hiroshi Tanaka", "Parent", "1970-09-12", "Male", "Japanese",
     "hiroshi.tanaka@example.jp", "+81 90 1234 9999", "Tokyo", "Japan", "NC-JP", "",
     "Father of Yuki · architect", "Active"],
    # Alumni
    ["CON-ALM-001", "Dr. Ingrid Olafsson", "Alumnus", "1990-06-18", "Female", "Norwegian",
     "ingrid.olafsson@who.int", "+41 22 791 2111", "Geneva", "Switzerland", "NC-NO", "",
     "UWC Red Cross Nordic '08 · WHO global health policy", "Active"],
    ["CON-ALM-002", "Prof. James Okonkwo", "Alumnus", "1977-02-25", "Male", "British-Nigerian",
     "j.okonkwo@ox.ac.uk", "+44 1865 270 000", "Oxford", "United Kingdom", "NC-UK", "",
     "UWC Atlantic '95 · climate scientist Oxford", "Active"],
    ["CON-ALM-003", "Anika Patel", "Alumna", "1997-04-09", "Female", "Indian",
     "anika@asha-edu.org", "+91 98765 22233", "Mumbai", "India", "NC-IN", "",
     "Pearson College UWC '15 · social entrepreneur (girls' education)", "Active"],
    ["CON-ALM-004", "Carlos Mendoza", "Alumnus", "1994-08-11", "Male", "Mexican-American",
     "c.mendoza@npr.org", "+1 202 513 2000", "Washington DC", "United States", "NC-US", "",
     "UWC-USA '12 · NPR international correspondent", "Active"],
    # Selectors / Volunteers
    ["CON-SEL-001", "Dr. Susanne Becker", "Selector", "1968-04-20", "Female", "German",
     "susanne.becker@volunteer.de.uwc.org", "+49 69 555 7777", "Frankfurt", "Germany", "NC-DE", "",
     "NC Selector — German NC · pediatrician · 8 yrs selection committee", "Active"],
    ["CON-SEL-002", "João Vieira", "Selector", "1975-12-03", "Male", "Brazilian",
     "joao.vieira@volunteer.br.uwc.org", "+55 11 9 5555 0001", "São Paulo", "Brazil", "NC-BR", "",
     "NC Selector — Brazil NC · academic dean", "Active"],
    ["CON-SEL-003", "Margaret Wanjiku", "Selector", "1972-08-08", "Female", "Kenyan",
     "margaret.wanjiku@volunteer.ke.uwc.org", "+254 722 555 999", "Nairobi", "Kenya", "NC-KE", "",
     "NC Selector — Kenya NC · UWC Atlantic '92 alumna", "Active"],
    ["CON-SEL-004", "Clara Ramos", "Selector", "1981-11-15", "Female", "Brazilian",
     "clara.ramos@volunteer.br.uwc.org", "+55 11 9 5555 0002", "São Paulo", "Brazil", "NC-BR", "",
     "NC Reviewer — Brazil · UWC Adriatic '99 · architect", "Active"],
    ["CON-SEL-005", "Paulo Fonseca", "Selector", "1980-02-26", "Male", "Brazilian",
     "paulo.fonseca@volunteer.br.uwc.org", "+55 11 9 5555 0003", "São Paulo", "Brazil", "NC-BR", "",
     "NC Reviewer — Brazil · UWC Mostar '98 · educator", "Active"],
    # Donors (Contact-level)
    ["CON-DON-001", "Shelby M.C. Davis", "Donor", "1937-04-29", "Male", "American",
     "office@davisfoundation.org", "+1 802 658 8200", "Charlotte VT", "United States", "NC-US", "ACC-DON-002",
     "Founder of Davis Scholars Programme · primary individual donor", "Active"],
    ["CON-DON-002", "Helmut Weber", "Donor", "1955-09-14", "Male", "German",
     "h.weber@weber-stiftung.de", "+49 30 555 1111", "Berlin", "Germany", "NC-DE", "ACC-DON-003",
     "Family foundation, €2.5M / yr to UWC Robert Bosch College", "Active"],
    ["CON-DON-003", "Brigitte Weber", "Donor", "1957-11-30", "Female", "German",
     "b.weber@weber-stiftung.de", "+49 30 555 1112", "Berlin", "Germany", "NC-DE", "ACC-DON-003",
     "Co-trustee, Weber Foundation", "Active"],
    ["CON-DON-004", "Anand Mahindra", "Donor", "1955-05-01", "Male", "Indian",
     "anand@mahindra.com", "+91 22 2490 1441", "Mumbai", "India", "NC-IN", "ACC-DON-004",
     "Chairman, Mahindra Group · primary donor UWC Mahindra College", "Active"],
]
write_sheet(wb, "04_Contacts", con_headers, con_rows,
            [13, 24, 11, 13, 9, 22, 30, 19, 16, 16, 10, 15, 44, 9])


# ============================================================
# SHEET 5: Leads (12)
# ============================================================
lead_headers = ["Lead ID", "Lead Name", "Email", "Phone", "Country", "Lead Source",
                "Interested Programme", "Owning NC", "Stage", "Lead Score",
                "Created Date", "Owner"]
lead_rows = [
    ["LEAD-001", "Henrik Bauer", "henrik.bauer@noemail.invalid", "+49 89 555 8801", "Germany", "Website",
     "UWC Robert Bosch College", "NC-DE", "New", 68, "2026-04-12", "Marcus Weber"],
    ["LEAD-002", "Lucia Fonseca", "lucia.fonseca@noemail.invalid", "+55 11 9 8801 0002", "Brazil", "School Visit",
     "UWC Costa Rica", "NC-BR", "Qualified", 78, "2026-04-15", "Ana Carvalho"],
    ["LEAD-003", "Arjun Reddy", "arjun.reddy@noemail.invalid", "+91 98765 88003", "India", "Open Day",
     "UWC Mahindra College", "NC-IN", "Contacted", 72, "2026-04-18", "Vikram Mehta"],
    ["LEAD-004", "Esther Kimani", "esther.kimani@noemail.invalid", "+254 722 800 004", "Kenya", "Alumni Referral",
     "UWC East Africa", "NC-KE", "Qualified", 85, "2026-04-21", "Joyce Wambui"],
    ["LEAD-005", "Tobias Lindqvist", "tobias.lindqvist@noemail.invalid", "+46 70 800 0005", "Sweden", "Webinar",
     "UWC Red Cross Nordic", "NC-NO", "New", 60, "2026-04-22", "Ingrid Olafsson"],
    ["LEAD-006", "Mei Lin Wong", "meilin.wong@noemail.invalid", "+852 9876 0006", "Hong Kong", "Social Media",
     "Li Po Chun UWC", "NC-IN", "Contacted", 70, "2026-04-25", "Vikram Mehta"],
    ["LEAD-007", "Carolina Vega", "carolina.vega@noemail.invalid", "+54 11 5555 0007", "Argentina", "School Counsellor",
     "UWC Adriatic", "NC-AR", "Qualified", 81, "2026-04-28", "Mariana Acosta"],
    ["LEAD-008", "Olivia Brown", "olivia.brown@noemail.invalid", "+44 7700 800 008", "United Kingdom", "Open Day",
     "UWC Atlantic", "NC-UK", "Disqualified", 35, "2026-04-30", "James Whitfield OBE"],
    ["LEAD-009", "Hiroto Watanabe", "hiroto.watanabe@noemail.invalid", "+81 90 8000 0009", "Japan", "Newsletter",
     "UWC ISAK Japan", "NC-JP", "Qualified", 88, "2026-05-02", "Hiroshi Tanaka"],
    ["LEAD-010", "Emma Dubois", "emma.dubois@noemail.invalid", "+33 6 8000 0010", "France", "Website",
     "UWC Atlantic", "NC-FR", "New", 65, "2026-05-05", "Pierre Lefebvre"],
    ["LEAD-011", "Daniel Park", "daniel.park@noemail.invalid", "+1 415 800 0011", "United States", "Davis Programme",
     "UWC-USA", "NC-US", "Converted", 92, "2026-05-08", "Sarah O'Brien"],
    ["LEAD-012", "Zara Ali", "zara.ali@noemail.invalid", "+44 7700 800 012", "United Kingdom", "Friend Referral",
     "UWC Maastricht", "NC-UK", "Qualified", 76, "2026-05-10", "James Whitfield OBE"],
]
write_sheet(wb, "05_Leads", lead_headers, lead_rows,
            [10, 22, 32, 18, 16, 18, 32, 10, 14, 11, 13, 22])


# ============================================================
# SHEET 6: Applications (13) — the pipeline centerpiece
# ============================================================
app_headers = ["Application ID", "Application Name", "Applicant Contact ID", "Owning NC",
               "Interested Programmes", "Selected Programme", "Stage", "Probability %",
               "Amount (GBP)", "Submission Date", "Closing Date", "Eligibility Status",
               "Application Language", "Owner"]
app_rows = [
    ["APP-2026-00073", "Lena Fischer Application 2026", "CON-STU-001", "NC-DE",
     "PRG-001; PRG-002", "—", "Under Review", 60,
     45000, "2026-05-25", "2026-06-15", "Eligible",
     "Deutsch", "Marcus Weber"],
    ["APP-2026-00074", "Klaus Berger Application 2026", "CON-STU-002", "NC-DE",
     "PRG-012", "PRG-012", "Nominated", 80,
     78000, "2026-05-18", "2026-06-12", "Eligible",
     "Deutsch", "Marcus Weber"],
    ["APP-2026-00075", "Anneke Müller Application 2026", "CON-STU-003", "NC-DE",
     "PRG-007; PRG-001", "—", "Shortlisted", 60,
     48000, "2026-05-20", "2026-06-12", "Eligible",
     "Deutsch", "Marcus Weber"],
    ["APP-2026-00076", "Amara Diallo Application 2026", "CON-STU-004", "NC-DE",
     "PRG-005; PRG-002", "PRG-005", "Placed", 100,
     56000, "2026-04-10", "2026-05-30", "Eligible",
     "Deutsch", "Marcus Weber"],
    ["APP-2026-00041", "Sofia Almeida Application 2026", "CON-STU-005", "NC-BR",
     "PRG-003; PRG-002", "—", "Submitted", 40,
     42000, "2026-05-22", "2026-06-20", "Eligible",
     "Português", "Ana Carvalho"],
    ["APP-2026-00042", "Rafael Mendes Application 2026", "CON-STU-006", "NC-BR",
     "PRG-016", "PRG-016", "Nominated", 80,
     92000, "2026-05-12", "2026-06-15", "Eligible",
     "Português", "Ana Carvalho"],
    ["APP-2026-00088", "Aarav Sharma Application 2026", "CON-STU-007", "NC-IN",
     "PRG-008", "PRG-008", "Under Review", 60,
     50000, "2026-05-19", "2026-06-18", "Eligible",
     "English", "Vikram Mehta"],
    ["APP-2026-00089", "Priya Iyer Application 2026", "CON-STU-008", "NC-IN",
     "PRG-006; PRG-008", "—", "In Progress", 20,
     50000, "—", "2026-06-25", "Pending",
     "English", "Vikram Mehta"],
    ["APP-2026-00031", "Kamau Njoroge Application 2026", "CON-STU-009", "NC-KE",
     "PRG-005; PRG-017", "PRG-005", "Placed", 100,
     56000, "2026-04-05", "2026-05-25", "Eligible",
     "English", "Joyce Wambui"],
    ["APP-2026-00112", "Yuki Tanaka Application 2026", "CON-STU-010", "NC-JP",
     "PRG-006", "PRG-006", "Shortlisted", 60,
     96000, "2026-05-15", "2026-06-20", "Eligible",
     "Japanese", "Hiroshi Tanaka"],
    ["APP-2026-00201", "Oliver Whitfield Application 2026", "CON-STU-011", "NC-UK",
     "PRG-001", "—", "Withdrawn", 0,
     90000, "2026-05-10", "2026-06-15", "Eligible",
     "English", "James Whitfield OBE"],
    ["APP-2026-00099", "Ayşe Demir Application 2026", "CON-STU-012", "NC-IT",
     "PRG-002; PRG-009", "—", "Declined", 0,
     72000, "2026-04-20", "2026-05-25", "Ineligible",
     "Türkçe", "Giulia Romano"],
    ["APP-2026-00077", "Henrik Bauer Application 2026", "—", "NC-DE",
     "PRG-012", "—", "Registered", 10,
     78000, "—", "2026-06-30", "Pending",
     "Deutsch", "Marcus Weber"],
]
write_sheet(wb, "06_Applications", app_headers, app_rows,
            [17, 30, 18, 9, 22, 18, 14, 11, 12, 13, 12, 14, 18, 22])


# ============================================================
# SHEET 7: Interviews (12)
# ============================================================
int_headers = ["Interview ID", "Application ID", "Applicant", "Interview Date",
               "Interview Mode", "Location", "Interviewer 1", "Interviewer 2",
               "Communication Score (/10)", "Leadership Score (/10)",
               "Final Recommendation", "Notes"]
int_rows = [
    ["INT-001", "APP-2026-00073", "Lena Fischer", "2026-05-20",
     "In-person", "UWC Bonn office", "CON-SEL-001", "CON-PAR-001",
     8, 9, "Recommend — Shortlist", "Articulate, strong English fluency"],
    ["INT-002", "APP-2026-00074", "Klaus Berger", "2026-05-15",
     "In-person", "UWC Berlin office", "CON-SEL-001", "—",
     9, 8, "Recommend — Nominate", "Excellent maths reasoning"],
    ["INT-003", "APP-2026-00075", "Anneke Müller", "2026-05-17",
     "Hybrid (in-person + Zoom)", "UWC Munich office", "CON-SEL-001", "—",
     8, 9, "Recommend — Shortlist", "Strong climate-action portfolio"],
    ["INT-004", "APP-2026-00076", "Amara Diallo", "2026-04-08",
     "In-person", "UWC Frankfurt office", "CON-SEL-001", "—",
     9, 10, "Recommend — Place", "Outstanding leadership"],
    ["INT-005", "APP-2026-00041", "Sofia Almeida", "2026-05-19",
     "In-person", "UWC São Paulo office", "CON-SEL-002", "CON-SEL-004",
     7, 9, "Recommend — Shortlist", "Genuine UWC mission alignment, modest English"],
    ["INT-006", "APP-2026-00042", "Rafael Mendes", "2026-05-09",
     "In-person", "UWC Rio de Janeiro office", "CON-SEL-002", "CON-SEL-005",
     9, 8, "Recommend — Nominate", "Strong technical maturity"],
    ["INT-007", "APP-2026-00088", "Aarav Sharma", "2026-05-17",
     "In-person", "UWC Mumbai office", "—", "—",
     8, 8, "Recommend — Shortlist", "Robotics leadership demonstrated"],
    ["INT-008", "APP-2026-00031", "Kamau Njoroge", "2026-04-03",
     "In-person", "UWC Nairobi office", "CON-SEL-003", "—",
     8, 10, "Recommend — Place", "Conservation activism authentic"],
    ["INT-009", "APP-2026-00112", "Yuki Tanaka", "2026-05-13",
     "In-person", "UWC Tokyo office", "—", "—",
     9, 9, "Recommend — Shortlist", "Strong written submission"],
    ["INT-010", "APP-2026-00201", "Oliver Whitfield", "2026-05-08",
     "In-person", "UWC London office", "—", "—",
     7, 6, "Decline (subsequently withdrew)", "Cambridge offer in hand"],
    ["INT-011", "APP-2026-00099", "Ayşe Demir", "2026-04-18",
     "Online", "Zoom", "—", "—",
     7, 8, "Decline (eligibility issue)", "Citizenship documentation incomplete"],
    ["INT-012", "APP-2026-00077", "Henrik Bauer", "2026-06-05",
     "Scheduled", "UWC Munich office", "CON-SEL-001", "—",
     0, 0, "Scheduled", "Interview pending"],
]
write_sheet(wb, "07_Interviews", int_headers, int_rows,
            [10, 17, 18, 13, 22, 24, 13, 13, 11, 11, 26, 36])


# ============================================================
# SHEET 8: Review_Scores (18 — 2 reviewers per app across 9 apps)
# ============================================================
rev_headers = ["Review ID", "Application ID", "Reviewer Contact ID", "Reviewer Name",
               "Blind?", "Academic Score (/10)", "Motivation Score (/10)",
               "Leadership Score (/10)", "Diversity Score (/10)", "Total (/40)",
               "Reviewer Comment", "Submitted Date"]
rev_rows = [
    ["REV-001", "APP-2026-00073", "CON-SEL-004", "Clara Ramos",
     "Yes", 9, 9, 8, 7, "=SUM(F2:I2)",
     "Articulate, modest financial means align with UWC mission", "2026-05-22"],
    ["REV-002", "APP-2026-00073", "CON-SEL-005", "Paulo Fonseca",
     "Yes", 8, 9, 9, 7, "=SUM(F3:I3)",
     "Strong leadership, recommend shortlist", "2026-05-22"],
    ["REV-003", "APP-2026-00074", "CON-SEL-001", "Dr. Susanne Becker",
     "No", 9, 8, 8, 6, "=SUM(F4:I4)",
     "Outstanding maths, recommend nomination", "2026-05-17"],
    ["REV-004", "APP-2026-00074", "CON-SEL-004", "Clara Ramos",
     "Yes", 9, 8, 7, 6, "=SUM(F5:I5)",
     "Confirm Robert Bosch College fit", "2026-05-17"],
    ["REV-005", "APP-2026-00075", "CON-SEL-001", "Dr. Susanne Becker",
     "No", 8, 9, 9, 8, "=SUM(F6:I6)",
     "Climate-action portfolio compelling", "2026-05-19"],
    ["REV-006", "APP-2026-00075", "CON-SEL-005", "Paulo Fonseca",
     "Yes", 8, 9, 8, 8, "=SUM(F7:I7)",
     "Recommend shortlist", "2026-05-19"],
    ["REV-007", "APP-2026-00076", "CON-SEL-001", "Dr. Susanne Becker",
     "No", 9, 10, 10, 9, "=SUM(F8:I8)",
     "Top of cohort. Place at UWC East Africa", "2026-04-12"],
    ["REV-008", "APP-2026-00076", "CON-SEL-003", "Margaret Wanjiku",
     "Yes", 9, 10, 10, 9, "=SUM(F9:I9)",
     "Outstanding across all dimensions", "2026-04-12"],
    ["REV-009", "APP-2026-00041", "CON-SEL-002", "João Vieira",
     "No", 8, 9, 9, 9, "=SUM(F10:I10)",
     "Strong public-school candidate", "2026-05-21"],
    ["REV-010", "APP-2026-00041", "CON-SEL-005", "Paulo Fonseca",
     "Yes", 7, 9, 9, 9, "=SUM(F11:I11)",
     "Recommend shortlist for UWC Costa Rica", "2026-05-21"],
    ["REV-011", "APP-2026-00042", "CON-SEL-004", "Clara Ramos",
     "Yes", 9, 8, 8, 7, "=SUM(F12:I12)",
     "Technical maturity strong, recommend nomination", "2026-05-11"],
    ["REV-012", "APP-2026-00042", "CON-SEL-005", "Paulo Fonseca",
     "Yes", 9, 8, 8, 7, "=SUM(F13:I13)",
     "Concur — nominate UWC-USA", "2026-05-11"],
    ["REV-013", "APP-2026-00088", "—", "—",
     "Yes", 0, 0, 0, 0, "=SUM(F14:I14)",
     "Pending reviewer assignment", "—"],
    ["REV-014", "APP-2026-00031", "CON-SEL-003", "Margaret Wanjiku",
     "No", 9, 10, 10, 9, "=SUM(F15:I15)",
     "Recommend immediate placement UWC East Africa", "2026-04-05"],
    ["REV-015", "APP-2026-00031", "CON-SEL-001", "Dr. Susanne Becker",
     "Yes", 9, 10, 10, 9, "=SUM(F16:I16)",
     "Concur — place", "2026-04-05"],
    ["REV-016", "APP-2026-00112", "—", "—",
     "Yes", 9, 9, 9, 8, "=SUM(F17:I17)",
     "Strong all-rounder", "2026-05-15"],
    ["REV-017", "APP-2026-00099", "—", "—",
     "Yes", 7, 8, 7, 7, "=SUM(F18:I18)",
     "Decline — eligibility issue (citizenship docs)", "2026-04-22"],
    ["REV-018", "APP-2026-00077", "—", "—",
     "Yes", 0, 0, 0, 0, "=SUM(F19:I19)",
     "Pending — application not yet submitted", "—"],
]
write_sheet(wb, "08_Review_Scores", rev_headers, rev_rows,
            [10, 17, 18, 22, 8, 11, 11, 11, 11, 11, 38, 13])


# ============================================================
# SHEET 9: Documents (18)
# ============================================================
doc_headers = ["Document ID", "Application ID", "Contact ID", "Document Type",
               "Filename", "Size (KB)", "Uploaded Date", "Uploaded By",
               "Status", "Verified By", "Verified Date"]
doc_rows = [
    ["DOC-001", "APP-2026-00073", "CON-STU-001", "Academic Transcript",
     "Fischer_L_Abitur_2026.pdf", 487, "2026-05-12", "CON-STU-001",
     "Verified", "Marcus Weber", "2026-05-13"],
    ["DOC-002", "APP-2026-00073", "CON-STU-001", "Passport Copy",
     "Fischer_L_Passport.pdf", 312, "2026-05-12", "CON-STU-001",
     "Verified", "Marcus Weber", "2026-05-13"],
    ["DOC-003", "APP-2026-00073", "CON-STU-001", "Recommendation Letter",
     "Fischer_L_Recommendation_Headmaster.pdf", 198, "2026-05-14", "ACC-SCH-001",
     "Received", "—", "—"],
    ["DOC-004", "APP-2026-00073", "CON-STU-001", "Personal Statement",
     "Fischer_L_Personal_Statement_DE.pdf", 142, "2026-05-15", "CON-STU-001",
     "Verified", "Marcus Weber", "2026-05-16"],
    ["DOC-005", "APP-2026-00074", "CON-STU-002", "Academic Transcript",
     "Berger_K_Transcript.pdf", 521, "2026-05-10", "CON-STU-002",
     "Verified", "Marcus Weber", "2026-05-11"],
    ["DOC-006", "APP-2026-00075", "CON-STU-003", "Academic Transcript",
     "Mueller_A_Transcript.pdf", 489, "2026-05-13", "CON-STU-003",
     "Verified", "Marcus Weber", "2026-05-14"],
    ["DOC-007", "APP-2026-00075", "CON-STU-003", "Climate Action Portfolio",
     "Mueller_A_Climate_Portfolio.pdf", 1842, "2026-05-13", "CON-STU-003",
     "Verified", "Marcus Weber", "2026-05-14"],
    ["DOC-008", "APP-2026-00076", "CON-STU-004", "Academic Transcript",
     "Diallo_A_Abitur.pdf", 502, "2026-04-05", "CON-STU-004",
     "Verified", "Marcus Weber", "2026-04-06"],
    ["DOC-009", "APP-2026-00076", "CON-STU-004", "MUN Participation Letter",
     "Diallo_A_MUN_Ambassador.pdf", 187, "2026-04-05", "CON-STU-004",
     "Verified", "Marcus Weber", "2026-04-06"],
    ["DOC-010", "APP-2026-00041", "CON-STU-005", "Academic Transcript",
     "Almeida_S_Transcript_PT.pdf", 376, "2026-05-18", "CON-STU-005",
     "Verified", "Ana Carvalho", "2026-05-19"],
    ["DOC-011", "APP-2026-00041", "CON-STU-005", "Income Statement (Mother)",
     "Almeida_M_Income_Statement.pdf", 124, "2026-05-19", "CON-PAR-002",
     "Verified", "Ana Carvalho", "2026-05-20"],
    ["DOC-012", "APP-2026-00041", "CON-STU-005", "Community Project Evidence",
     "Almeida_S_Favela_Project_Photos.pdf", 3214, "2026-05-19", "CON-STU-005",
     "Verified", "Ana Carvalho", "2026-05-20"],
    ["DOC-013", "APP-2026-00042", "CON-STU-006", "Academic Transcript",
     "Mendes_R_Transcript.pdf", 412, "2026-05-08", "CON-STU-006",
     "Verified", "Ana Carvalho", "2026-05-09"],
    ["DOC-014", "APP-2026-00088", "CON-STU-007", "Academic Transcript",
     "Sharma_A_ICSE_Transcript.pdf", 388, "2026-05-15", "CON-STU-007",
     "Received", "—", "—"],
    ["DOC-015", "APP-2026-00088", "CON-STU-007", "Robotics Competition Certificate",
     "Sharma_A_RoboCup_Certificate.pdf", 156, "2026-05-15", "CON-STU-007",
     "Received", "—", "—"],
    ["DOC-016", "APP-2026-00031", "CON-STU-009", "Academic Transcript",
     "Njoroge_K_Transcript.pdf", 405, "2026-04-01", "CON-STU-009",
     "Verified", "Joyce Wambui", "2026-04-02"],
    ["DOC-017", "APP-2026-00112", "CON-STU-010", "Academic Transcript",
     "Tanaka_Y_Transcript_JP.pdf", 522, "2026-05-11", "CON-STU-010",
     "Verified", "Hiroshi Tanaka", "2026-05-12"],
    ["DOC-018", "APP-2026-00099", "CON-STU-012", "Citizenship Documentation",
     "Demir_A_Citizenship_Issue.pdf", 287, "2026-04-15", "CON-STU-012",
     "Rejected", "Giulia Romano", "2026-04-18"],
]
write_sheet(wb, "09_Documents", doc_headers, doc_rows,
            [10, 17, 13, 24, 36, 11, 13, 17, 11, 18, 13])


# ============================================================
# SHEET 10: Consent_Records (12)
# ============================================================
cr_headers = ["Consent ID", "Contact ID", "Application ID", "Consent Type",
              "Granted?", "Granted Date", "Source", "Granted By",
              "Lawful Basis (GDPR Art.6)", "Special Category Basis (Art.9)",
              "Expires Date", "Notes"]
cr_rows = [
    ["CR-001", "CON-STU-001", "APP-2026-00073", "Parental Consent (Minor Applicant)",
     "Yes", "2026-05-24", "Application portal", "CON-PAR-001",
     "6(1)(a) Consent", "9(2)(a) Explicit Consent",
     "2028-05-30", "Signed by Petra Fischer (mother)"],
    ["CR-002", "CON-STU-001", "APP-2026-00073", "Safeguarding Disclosure",
     "Yes", "2026-05-24", "Application portal", "CON-PAR-001",
     "6(1)(a) Consent", "9(2)(a) Explicit Consent",
     "2028-05-30", "Standard UWC safeguarding declaration"],
    ["CR-003", "CON-STU-002", "APP-2026-00074", "Parental Consent (Minor Applicant)",
     "Yes", "2026-05-15", "Application portal", "—",
     "6(1)(a) Consent", "9(2)(a) Explicit Consent",
     "2028-05-30", "Both parents signed"],
    ["CR-004", "CON-STU-003", "APP-2026-00075", "Parental Consent (Minor Applicant)",
     "Yes", "2026-05-18", "Application portal", "—",
     "6(1)(a) Consent", "9(2)(a) Explicit Consent",
     "2028-05-30", "Mother sole guardian"],
    ["CR-005", "CON-STU-004", "APP-2026-00076", "Parental Consent (Minor Applicant)",
     "Yes", "2026-04-08", "In-person form", "—",
     "6(1)(a) Consent", "9(2)(a) Explicit Consent",
     "2028-05-30", "Signed at Frankfurt interview"],
    ["CR-006", "CON-STU-005", "APP-2026-00041", "Parental Consent (Minor Applicant)",
     "Yes", "2026-05-21", "Application portal", "CON-PAR-002",
     "6(1)(a) Consent", "9(2)(a) Explicit Consent",
     "2028-05-30", "Maria Almeida (mother)"],
    ["CR-007", "CON-STU-006", "APP-2026-00042", "Parental Consent (Minor Applicant)",
     "Yes", "2026-05-12", "Application portal", "—",
     "6(1)(a) Consent", "9(2)(a) Explicit Consent",
     "2028-05-30", "Both parents signed"],
    ["CR-008", "CON-STU-007", "APP-2026-00088", "Parental Consent (Minor Applicant)",
     "Yes", "2026-05-19", "Application portal", "CON-PAR-003",
     "6(1)(a) Consent", "9(2)(a) Explicit Consent",
     "2028-05-30", "Anand Sharma (father)"],
    ["CR-009", "CON-STU-009", "APP-2026-00031", "Parental Consent (Minor Applicant)",
     "Yes", "2026-04-04", "Application portal", "CON-PAR-004",
     "6(1)(a) Consent", "9(2)(a) Explicit Consent",
     "2028-05-30", "Grace Njoroge (mother)"],
    ["CR-010", "CON-STU-010", "APP-2026-00112", "Parental Consent (Minor Applicant)",
     "Yes", "2026-05-14", "Application portal", "CON-PAR-005",
     "6(1)(a) Consent", "9(2)(a) Explicit Consent",
     "2028-05-30", "Hiroshi Tanaka (father)"],
    ["CR-011", "CON-STU-008", "APP-2026-00089", "Parental Consent (Minor Applicant)",
     "Pending", "—", "Application portal", "—",
     "—", "—",
     "—", "Consent form not yet received"],
    ["CR-012", "CON-STU-012", "APP-2026-00099", "Marketing Consent (Future Comms)",
     "No", "2026-04-20", "Application portal", "CON-STU-012",
     "6(1)(a) Consent", "—",
     "—", "Applicant declined future marketing"],
]
write_sheet(wb, "10_Consent_Records", cr_headers, cr_rows,
            [10, 13, 17, 32, 10, 13, 18, 13, 24, 28, 12, 36])


# ============================================================
# SHEET 11: Financial_Aid (12)
# ============================================================
fa_headers = ["Financial Aid ID", "Application ID", "Applicant", "Aid Requested",
              "Household Income (Local)", "Currency", "Income (GBP equiv.)",
              "Aid Amount Requested (GBP)", "Aid Amount Approved (GBP)",
              "Scholarship Priority Score (/100)", "Financial Review Status", "Reviewer"]
fa_rows = [
    ["FA-001", "APP-2026-00073", "Lena Fischer", "Full Scholarship",
     "€42,000 / yr", "EUR", 36500, 45000, 0, 87,
     "Pending", "Marcus Weber"],
    ["FA-002", "APP-2026-00074", "Klaus Berger", "Partial (50%)",
     "€68,000 / yr", "EUR", 59000, 39000, 39000, 62,
     "Approved", "Marcus Weber"],
    ["FA-003", "APP-2026-00075", "Anneke Müller", "Partial (30%)",
     "€85,000 / yr", "EUR", 73800, 14400, 14400, 48,
     "Approved", "Marcus Weber"],
    ["FA-004", "APP-2026-00076", "Amara Diallo", "Full Scholarship",
     "€28,000 / yr", "EUR", 24300, 56000, 56000, 95,
     "Approved", "Marcus Weber"],
    ["FA-005", "APP-2026-00041", "Sofia Almeida", "Full Scholarship",
     "R$ 96,000 / yr", "BRL", 14500, 42000, 0, 91,
     "Pending", "Ana Carvalho"],
    ["FA-006", "APP-2026-00042", "Rafael Mendes", "Full Scholarship",
     "R$ 144,000 / yr", "BRL", 21800, 92000, 92000, 85,
     "Approved (Davis Programme)", "Ana Carvalho"],
    ["FA-007", "APP-2026-00088", "Aarav Sharma", "Partial (40%)",
     "₹ 36,00,000 / yr", "INR", 34000, 20000, 0, 65,
     "Pending", "Vikram Mehta"],
    ["FA-008", "APP-2026-00031", "Kamau Njoroge", "Full Scholarship",
     "KES 720,000 / yr", "KES", 4500, 56000, 56000, 97,
     "Approved", "Joyce Wambui"],
    ["FA-009", "APP-2026-00112", "Yuki Tanaka", "Partial (50%)",
     "¥ 9,200,000 / yr", "JPY", 48500, 48000, 0, 58,
     "Pending", "Hiroshi Tanaka"],
    ["FA-010", "APP-2026-00201", "Oliver Whitfield", "None Requested",
     "£180,000 / yr", "GBP", 180000, 0, 0, 0,
     "N/A — withdrawn", "James Whitfield OBE"],
    ["FA-011", "APP-2026-00099", "Ayşe Demir", "Full Scholarship",
     "₺ 380,000 / yr", "TRY", 9800, 72000, 0, 78,
     "Declined — Ineligible", "Giulia Romano"],
    ["FA-012", "APP-2026-00089", "Priya Iyer", "Partial (30%)",
     "₹ 48,00,000 / yr", "INR", 45500, 15000, 0, 52,
     "Not Submitted", "Vikram Mehta"],
]
write_sheet(wb, "11_Financial_Aid", fa_headers, fa_rows,
            [13, 17, 16, 16, 22, 9, 14, 16, 16, 14, 22, 19])


# ============================================================
# SHEET 12: Tasks (15)
# ============================================================
tsk_headers = ["Task ID", "Subject", "Related Record ID", "Related Module",
               "Owner", "Priority", "Due Date", "Status", "Notes"]
tsk_rows = [
    ["TSK-001", "Review application documents and verify eligibility", "APP-2026-00073", "Applications",
     "Marcus Weber", "High", "2026-05-28", "In Progress", "Today"],
    ["TSK-002", "Send confirmation email to applicant", "APP-2026-00073", "Applications",
     "Marcus Weber", "Medium", "2026-05-28", "Open", ""],
    ["TSK-003", "Schedule interview with Klaus Berger", "APP-2026-00074", "Applications",
     "Marcus Weber", "High", "2026-06-01", "Open", "Coordinate with selector"],
    ["TSK-004", "Verify financial documents — Anneke Müller", "APP-2026-00075", "Applications",
     "Marcus Weber", "Medium", "2026-06-02", "Open", ""],
    ["TSK-005", "Send placement confirmation — Amara Diallo to UWC East Africa", "APP-2026-00076", "Applications",
     "Marcus Weber", "High", "2026-05-30", "Open", "Coordinate with Kenya NC"],
    ["TSK-006", "Review Sofia Almeida portfolio", "APP-2026-00041", "Applications",
     "Ana Carvalho", "High", "2026-05-29", "In Progress", ""],
    ["TSK-007", "Approve full scholarship — Rafael Mendes (Davis Programme)", "APP-2026-00042", "Applications",
     "Ana Carvalho", "High", "2026-05-30", "Completed", ""],
    ["TSK-008", "Assign 2 reviewers to Aarav Sharma application", "APP-2026-00088", "Applications",
     "Vikram Mehta", "High", "2026-05-29", "Open", ""],
    ["TSK-009", "Follow up on missing transcript — Priya Iyer", "APP-2026-00089", "Applications",
     "Vikram Mehta", "Medium", "2026-06-05", "Open", ""],
    ["TSK-010", "Send placement onboarding pack — Kamau Njoroge", "APP-2026-00031", "Applications",
     "Joyce Wambui", "Medium", "2026-06-01", "Open", "Include visa guidance"],
    ["TSK-011", "Schedule final interview — Yuki Tanaka", "APP-2026-00112", "Applications",
     "Hiroshi Tanaka", "Medium", "2026-06-03", "Open", ""],
    ["TSK-012", "Process withdrawal — Oliver Whitfield", "APP-2026-00201", "Applications",
     "James Whitfield OBE", "Low", "2026-05-30", "Open", "Send confirmation email"],
    ["TSK-013", "GDPR — close out Ayşe Demir declined application", "APP-2026-00099", "Applications",
     "Giulia Romano", "Medium", "2026-06-01", "Open", "30-day data retention notice"],
    ["TSK-014", "Quarterly NC report — UWC Germany", "NC-DE", "National_Committees",
     "Marcus Weber", "Medium", "2026-06-30", "Open", "Q2 statistics"],
    ["TSK-015", "Selector training refresh — Susanne Becker", "CON-SEL-001", "Contacts",
     "Marcus Weber", "Low", "2026-09-15", "Open", "Annual GDPR + safeguarding"],
]
write_sheet(wb, "12_Tasks", tsk_headers, tsk_rows,
            [10, 48, 17, 18, 22, 10, 12, 13, 22])


# ============================================================
# SHEET 13: Meetings (12)
# ============================================================
mtg_headers = ["Meeting ID", "Subject", "Meeting Type", "Related Record ID",
               "Related Module", "Start Date/Time", "End Date/Time", "Location",
               "Organizer", "Attendees", "Status"]
mtg_rows = [
    ["MTG-001", "Selection Committee — DE Cohort Round 1", "Selection Committee", "PRG-CMP-DE26",
     "Programmes", "2026-05-30 09:00", "2026-05-30 13:00", "UWC Bonn office",
     "Marcus Weber", "Susanne Becker; Helen Richards; 4 reviewers", "Scheduled"],
    ["MTG-002", "Lena Fischer Interview", "Interview", "APP-2026-00073",
     "Applications", "2026-05-20 14:00", "2026-05-20 15:30", "UWC Bonn office",
     "Marcus Weber", "Lena Fischer; Susanne Becker", "Completed"],
    ["MTG-003", "Klaus Berger Interview", "Interview", "APP-2026-00074",
     "Applications", "2026-05-15 10:00", "2026-05-15 11:30", "UWC Berlin office",
     "Marcus Weber", "Klaus Berger; Susanne Becker", "Completed"],
    ["MTG-004", "Sofia Almeida Interview", "Interview", "APP-2026-00041",
     "Applications", "2026-05-19 15:00", "2026-05-19 16:30", "UWC São Paulo office",
     "Ana Carvalho", "Sofia Almeida; João Vieira; Clara Ramos", "Completed"],
    ["MTG-005", "BR Selection Committee — Round 1", "Selection Committee", "PRG-CMP-BR26",
     "Programmes", "2026-06-02 09:00", "2026-06-02 13:00", "UWC São Paulo office",
     "Ana Carvalho", "João Vieira; Clara Ramos; Paulo Fonseca", "Scheduled"],
    ["MTG-006", "NC Quarterly Governance — DE", "Governance", "NC-DE",
     "National_Committees", "2026-06-15 10:00", "2026-06-15 12:00", "Bonn HQ + Zoom",
     "Marcus Weber", "All NC-DE selectors + IO observer", "Scheduled"],
    ["MTG-007", "Donor Stewardship — Bosch Stiftung", "Donor Meeting", "ACC-DON-003",
     "Accounts", "2026-06-10 14:00", "2026-06-10 15:00", "Robert Bosch Stiftung, Stuttgart",
     "Marcus Weber", "Bernhard Straub; Helmut Weber", "Scheduled"],
    ["MTG-008", "Kamau Njoroge Placement Briefing", "Placement", "APP-2026-00031",
     "Applications", "2026-06-01 11:00", "2026-06-01 12:00", "Zoom",
     "Joyce Wambui", "Kamau Njoroge; Grace Njoroge; UWC EAfrica admissions", "Scheduled"],
    ["MTG-009", "Aarav Sharma Interview", "Interview", "APP-2026-00088",
     "Applications", "2026-05-17 10:00", "2026-05-17 11:30", "UWC Mumbai office",
     "Vikram Mehta", "Aarav Sharma; Anand Sharma; 2 reviewers", "Completed"],
    ["MTG-010", "IO Senior Leadership — Selection Round Up", "IO Leadership", "NC-UK",
     "National_Committees", "2026-07-05 14:00", "2026-07-05 16:00", "UWC International HQ London",
     "James Whitfield OBE", "All NC Directors; IO Admissions team", "Scheduled"],
    ["MTG-011", "Yuki Tanaka Interview", "Interview", "APP-2026-00112",
     "Applications", "2026-05-13 13:00", "2026-05-13 14:30", "UWC Tokyo office",
     "Hiroshi Tanaka", "Yuki Tanaka; Hiroshi Tanaka (NC Director); 1 reviewer", "Completed"],
    ["MTG-012", "Davis Scholars Coordination — US Round", "Donor Meeting", "ACC-DON-001",
     "Accounts", "2026-06-12 16:00", "2026-06-12 17:00", "Zoom",
     "Sarah O'Brien", "Philip Geier Jr.; Sarah O'Brien; 3 NC Directors", "Scheduled"],
]
write_sheet(wb, "13_Meetings", mtg_headers, mtg_rows,
            [10, 36, 18, 17, 18, 18, 18, 28, 20, 36, 11])


# ============================================================
# SHEET 14: Calls (12)
# ============================================================
call_headers = ["Call ID", "Subject", "Call Type", "Related Record ID", "Related Module",
                "Caller (Internal)", "Recipient Contact", "Call Date/Time", "Duration (min)",
                "Outcome", "Notes"]
call_rows = [
    ["CALL-001", "Interview confirmation — Lena Fischer", "Outbound", "APP-2026-00073", "Applications",
     "Marcus Weber", "CON-PAR-001", "2026-05-19 10:30", 8,
     "Confirmed", "Petra Fischer confirmed attendance"],
    ["CALL-002", "Document follow-up — Anneke Müller", "Outbound", "APP-2026-00075", "Applications",
     "Marcus Weber", "CON-STU-003", "2026-05-17 14:15", 6,
     "Action Required", "Asked for climate-action portfolio update"],
    ["CALL-003", "Placement confirmation — Amara Diallo", "Outbound", "APP-2026-00076", "Applications",
     "Marcus Weber", "CON-STU-004", "2026-04-15 16:00", 12,
     "Placed", "Confirmed UWC East Africa placement, briefed on visa"],
    ["CALL-004", "Income verification — Sofia Almeida", "Outbound", "APP-2026-00041", "Applications",
     "Ana Carvalho", "CON-PAR-002", "2026-05-20 11:00", 18,
     "Verified", "Maria confirmed household income, documents sent"],
    ["CALL-005", "Davis Programme briefing — Rafael Mendes", "Outbound", "APP-2026-00042", "Applications",
     "Ana Carvalho", "CON-STU-006", "2026-05-13 15:30", 22,
     "Briefed", "Explained Davis Programme + UWC-USA pathway"],
    ["CALL-006", "Missing transcript follow-up — Priya Iyer", "Outbound", "APP-2026-00089", "Applications",
     "Vikram Mehta", "CON-STU-008", "2026-05-22 10:00", 5,
     "Action Required", "Transcript to be sent by 28 May"],
    ["CALL-007", "Placement onboarding — Kamau Njoroge", "Outbound", "APP-2026-00031", "Applications",
     "Joyce Wambui", "CON-PAR-004", "2026-04-20 17:00", 25,
     "Onboarded", "Visa process briefed, financial aid confirmed"],
    ["CALL-008", "Parent enquiry — Yuki Tanaka", "Inbound", "APP-2026-00112", "Applications",
     "Hiroshi Tanaka", "CON-PAR-005", "2026-05-14 09:30", 14,
     "Information Provided", "Explained selection timeline"],
    ["CALL-009", "Withdrawal processing — Oliver Whitfield", "Inbound", "APP-2026-00201", "Applications",
     "James Whitfield OBE", "CON-STU-011", "2026-05-25 11:00", 7,
     "Withdrawn", "Cambridge offer accepted, formal withdrawal received"],
    ["CALL-010", "Ineligibility notification — Ayşe Demir", "Outbound", "APP-2026-00099", "Applications",
     "Giulia Romano", "CON-STU-012", "2026-04-22 14:00", 11,
     "Closed Negative", "Citizenship docs insufficient, recommended next year"],
    ["CALL-011", "Initial enquiry — Henrik Bauer", "Inbound", "LEAD-001", "Leads",
     "Marcus Weber", "—", "2026-04-12 16:00", 9,
     "Qualified", "Strong fit, encouraged to apply"],
    ["CALL-012", "Donor relationship — Helmut Weber", "Outbound", "ACC-DON-003", "Accounts",
     "Marcus Weber", "CON-DON-002", "2026-05-15 11:00", 28,
     "Stewardship", "Discussed Q3 disbursement to UWC Robert Bosch College"],
]
write_sheet(wb, "14_Calls", call_headers, call_rows,
            [10, 42, 11, 18, 18, 22, 22, 18, 13, 22, 36])


# ============================================================
# SHEET 15: Campaigns (10)
# ============================================================
cmp_headers = ["Campaign ID", "Campaign Name", "Type", "Status", "Owning NC",
               "Start Date", "End Date", "Budget (GBP)", "Expected Leads",
               "Actual Leads", "Conversions"]
cmp_rows = [
    ["CMP-001", "DE 2026 Selection — Spring Outreach", "Selection Outreach", "Active", "NC-DE",
     "2025-10-01", "2026-04-30", 18000, 200, 187, 14],
    ["CMP-002", "BR 2026 Selection — São Paulo + Rio", "Selection Outreach", "Active", "NC-BR",
     "2025-10-01", "2026-05-15", 14000, 180, 142, 12],
    ["CMP-003", "IN 2026 Selection — 3-City Tour", "Selection Outreach", "Active", "NC-IN",
     "2025-10-15", "2026-05-30", 22000, 350, 312, 28],
    ["CMP-004", "KE 2026 Selection — Nairobi Schools", "Selection Outreach", "Active", "NC-KE",
     "2025-11-01", "2026-04-30", 8000, 80, 75, 11],
    ["CMP-005", "UK 2026 Selection — London + Regional", "Selection Outreach", "Active", "NC-UK",
     "2025-09-15", "2026-05-15", 26000, 280, 245, 22],
    ["CMP-006", "US 2026 Davis Programme Outreach", "Donor-Funded Outreach", "Active", "NC-US",
     "2025-10-01", "2026-06-30", 38000, 400, 372, 35],
    ["CMP-007", "JP 2026 Senbatsu — Tokyo", "Selection Outreach", "Active", "NC-JP",
     "2025-11-15", "2026-04-30", 12000, 100, 87, 14],
    ["CMP-008", "Davis Scholars Donor Stewardship", "Donor Relations", "Active", "NC-US",
     "2026-01-01", "2026-12-31", 15000, 0, 0, 0],
    ["CMP-009", "Alumni Network Re-engagement", "Alumni Relations", "Active", "NC-UK",
     "2026-02-01", "2026-08-31", 9000, 0, 0, 0],
    ["CMP-010", "Climate Action UWC — Global Webinar Series", "Brand Awareness", "Planning", "NC-UK",
     "2026-09-01", "2026-12-15", 22000, 500, 0, 0],
]
write_sheet(wb, "15_Campaigns", cmp_headers, cmp_rows,
            [10, 42, 22, 10, 10, 12, 12, 13, 14, 13, 13])


# ============================================================
# SHEET 16: Stage_History (18)
# ============================================================
sh_headers = ["History ID", "Application ID", "From Stage", "To Stage", "Changed By",
              "Changed Date/Time", "Reason / Comment"]
sh_rows = [
    ["SH-001", "APP-2026-00073", "—", "Registered", "Marcus Weber", "2026-04-25 09:00", "Application created"],
    ["SH-002", "APP-2026-00073", "Registered", "In Progress", "CON-STU-001", "2026-05-12 11:30", "Applicant started form"],
    ["SH-003", "APP-2026-00073", "In Progress", "Submitted", "CON-STU-001", "2026-05-25 22:14", "Applicant submitted"],
    ["SH-004", "APP-2026-00073", "Submitted", "Under Review", "Marcus Weber", "2026-05-26 09:45", "Eligibility verified"],
    ["SH-005", "APP-2026-00074", "Submitted", "Under Review", "Marcus Weber", "2026-05-19 09:00", "Eligibility verified"],
    ["SH-006", "APP-2026-00074", "Under Review", "Shortlisted", "Marcus Weber", "2026-05-22 14:00", "Strong reviewer scores"],
    ["SH-007", "APP-2026-00074", "Shortlisted", "Nominated", "Marcus Weber", "2026-05-25 10:00", "Nominated to UWC Robert Bosch College"],
    ["SH-008", "APP-2026-00075", "Submitted", "Under Review", "Marcus Weber", "2026-05-21 09:30", "Eligibility verified"],
    ["SH-009", "APP-2026-00075", "Under Review", "Shortlisted", "Marcus Weber", "2026-05-23 16:00", "Recommended by 2 reviewers"],
    ["SH-010", "APP-2026-00076", "Shortlisted", "Nominated", "Marcus Weber", "2026-04-18 11:00", "Nominated to UWC East Africa"],
    ["SH-011", "APP-2026-00076", "Nominated", "Placed", "Joyce Wambui", "2026-04-22 14:30", "Place confirmed by EA admissions"],
    ["SH-012", "APP-2026-00041", "In Progress", "Submitted", "CON-STU-005", "2026-05-22 19:00", "Applicant submitted"],
    ["SH-013", "APP-2026-00042", "Submitted", "Nominated", "Ana Carvalho", "2026-05-18 11:00", "Davis Programme fast-track"],
    ["SH-014", "APP-2026-00088", "Submitted", "Under Review", "Vikram Mehta", "2026-05-20 10:00", "Eligibility verified"],
    ["SH-015", "APP-2026-00031", "Nominated", "Placed", "Joyce Wambui", "2026-04-15 09:00", "EA placement confirmed"],
    ["SH-016", "APP-2026-00112", "Submitted", "Shortlisted", "Hiroshi Tanaka", "2026-05-16 11:00", "Strong written + interview"],
    ["SH-017", "APP-2026-00201", "Under Review", "Withdrawn", "James Whitfield OBE", "2026-05-25 11:30", "Applicant accepted Cambridge offer"],
    ["SH-018", "APP-2026-00099", "Submitted", "Declined", "Giulia Romano", "2026-04-22 15:00", "Ineligible — citizenship docs"],
]
write_sheet(wb, "16_Stage_History", sh_headers, sh_rows,
            [10, 17, 15, 15, 22, 18, 36])


# ============================================================
# SHEET 17: Training_Records (12)
# ============================================================
trn_headers = ["Training ID", "Selector Contact ID", "Selector Name",
               "Training Module", "Completion Date", "Score (/100)", "Expires Date",
               "Certified By", "Status"]
trn_rows = [
    ["TRN-001", "CON-SEL-001", "Dr. Susanne Becker",
     "GDPR for NC Selectors (Annual)", "2026-01-15", 96, "2027-01-15", "IO Compliance", "Current"],
    ["TRN-002", "CON-SEL-001", "Dr. Susanne Becker",
     "Safeguarding Level 2 (Annual)", "2026-01-15", 94, "2027-01-15", "IO Safeguarding", "Current"],
    ["TRN-003", "CON-SEL-001", "Dr. Susanne Becker",
     "Unconscious Bias in Selection", "2026-02-08", 88, "2027-02-08", "IO Equity Team", "Current"],
    ["TRN-004", "CON-SEL-002", "João Vieira",
     "GDPR for NC Selectors (Annual)", "2026-02-12", 92, "2027-02-12", "IO Compliance", "Current"],
    ["TRN-005", "CON-SEL-002", "João Vieira",
     "Safeguarding Level 2 (Annual)", "2026-02-12", 89, "2027-02-12", "IO Safeguarding", "Current"],
    ["TRN-006", "CON-SEL-003", "Margaret Wanjiku",
     "GDPR for NC Selectors (Annual)", "2026-01-20", 95, "2027-01-20", "IO Compliance", "Current"],
    ["TRN-007", "CON-SEL-003", "Margaret Wanjiku",
     "Safeguarding Level 2 (Annual)", "2026-01-20", 91, "2027-01-20", "IO Safeguarding", "Current"],
    ["TRN-008", "CON-SEL-004", "Clara Ramos",
     "Blind Review Methodology", "2026-03-04", 87, "2027-03-04", "IO Selection Team", "Current"],
    ["TRN-009", "CON-SEL-004", "Clara Ramos",
     "Safeguarding Level 2 (Annual)", "2026-03-04", 90, "2027-03-04", "IO Safeguarding", "Current"],
    ["TRN-010", "CON-SEL-005", "Paulo Fonseca",
     "Blind Review Methodology", "2026-03-04", 89, "2027-03-04", "IO Selection Team", "Current"],
    ["TRN-011", "CON-SEL-005", "Paulo Fonseca",
     "GDPR for NC Selectors (Annual)", "2026-03-04", 93, "2027-03-04", "IO Compliance", "Current"],
    ["TRN-012", "CON-SEL-005", "Paulo Fonseca",
     "Unconscious Bias in Selection", "2025-09-12", 84, "2026-09-12", "IO Equity Team", "Expiring Soon"],
]
write_sheet(wb, "17_Training_Records", trn_headers, trn_rows,
            [10, 17, 22, 36, 14, 11, 12, 22, 14])


# ============================================================
# SHEET 18: Audit_Log (15)
# ============================================================
aud_headers = ["Audit ID", "Timestamp", "Actor (User)", "Action",
               "Affected Record ID", "Affected Module", "Detail / Field Changed",
               "Old Value", "New Value", "IP Address"]
aud_rows = [
    ["AUD-001", "2026-05-25 22:14:32", "CON-STU-001", "Submit Application",
     "APP-2026-00073", "Applications", "Stage", "In Progress", "Submitted", "92.45.220.18"],
    ["AUD-002", "2026-05-26 09:45:11", "Marcus Weber", "Verify Eligibility",
     "APP-2026-00073", "Applications", "Eligibility Status", "Pending", "Eligible", "194.245.45.10"],
    ["AUD-003", "2026-05-26 10:00:24", "Marcus Weber", "Change Stage",
     "APP-2026-00073", "Applications", "Stage", "Submitted", "Under Review", "194.245.45.10"],
    ["AUD-004", "2026-05-26 14:15:08", "Data Protection Lead", "View Sensitive Record",
     "APP-2026-00073", "Applications", "Read — Consent + Safeguarding section", "—", "Read", "194.245.45.40"],
    ["AUD-005", "2026-05-26 16:32:55", "Marcus Weber", "Unlock Application",
     "APP-2026-00078", "Applications", "Stage", "Submitted", "In Progress",
     "194.245.45.10"],
    ["AUD-006", "2026-05-22 14:00:12", "Marcus Weber", "Shortlist Application",
     "APP-2026-00074", "Applications", "Stage", "Under Review", "Shortlisted", "194.245.45.10"],
    ["AUD-007", "2026-05-25 10:00:08", "Marcus Weber", "Nominate Application",
     "APP-2026-00074", "Applications", "Stage", "Shortlisted", "Nominated", "194.245.45.10"],
    ["AUD-008", "2026-04-22 14:30:44", "Joyce Wambui", "Place Application",
     "APP-2026-00076", "Applications", "Stage", "Nominated", "Placed", "41.90.74.22"],
    ["AUD-009", "2026-05-22 19:00:55", "CON-STU-005", "Submit Application",
     "APP-2026-00041", "Applications", "Stage", "In Progress", "Submitted", "189.18.45.99"],
    ["AUD-010", "2026-05-18 11:00:30", "Ana Carvalho", "Nominate Application",
     "APP-2026-00042", "Applications", "Stage", "Submitted", "Nominated", "189.18.45.50"],
    ["AUD-011", "2026-05-25 11:30:21", "James Whitfield OBE", "Withdraw Application",
     "APP-2026-00201", "Applications", "Stage", "Under Review", "Withdrawn", "194.83.92.4"],
    ["AUD-012", "2026-04-22 15:00:08", "Giulia Romano", "Decline Application",
     "APP-2026-00099", "Applications", "Stage", "Submitted", "Declined", "151.45.78.22"],
    ["AUD-013", "2026-05-20 14:00:00", "System", "GDPR Consent Recorded",
     "CR-001", "Consent_Records", "Granted", "—", "Yes", "—"],
    ["AUD-014", "2026-05-28 09:12:30", "System", "GDPR Retention — Auto-purge scheduled",
     "APP-2026-00099", "Applications", "Retention scheduled (declined applicant)",
     "Active", "Purge 2027-04-22", "—"],
    ["AUD-015", "2026-05-26 14:00:00", "Akanksha Anand", "Role Assignment Change",
     "CON-SEL-005", "Contacts", "Role", "Reviewer", "Reviewer + Selector",
     "194.83.92.10"],
]
write_sheet(wb, "18_Audit_Log", aud_headers, aud_rows,
            [10, 20, 22, 30, 17, 18, 36, 15, 22, 14])


# ============================================================
# Place README first
# ============================================================
wb.move_sheet("README", offset=-(len(wb.sheetnames) - 1))

# ============================================================
# Save
# ============================================================
out_paths = [
    "/sessions/confident-pensive-pascal/mnt/workspace/uwc-demo-project/uwc-demo/data/UWC_CRM_Sample_Dataset.xlsx",
    "/sessions/confident-pensive-pascal/mnt/outputs/UWC_CRM_Sample_Dataset.xlsx",
]
for p in out_paths:
    wb.save(p)
    print(f"Saved: {p}")

# Sheet inventory
print()
print("Sheet inventory:")
for name in wb.sheetnames:
    ws = wb[name]
    print(f"  {name}: {ws.max_row} rows × {ws.max_column} cols")
